import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import mysql.connector
import pandas as pd
import os
import io
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

carrera_map = {
    "COMUNICACIÓN SOCIAL": "01",
    "EDUCACIÓN SOCIAL": "02",
    "EDUCACIÓN PRIMARIA": "03",
    "EDUCACIÓN SECUNDARIA": "04",
    "PSICOLOGÍA": "05",
    "PUBLICIDAD Y MULTIMEDIA": "06",
    "TEOLOGÍA": "07",
    "TRABAJO SOCIAL": "08",
    "TURISMO Y HOTELERÍA": "09",
    "ADMINISTRACIÓN DE EMPRESAS": "10",
    "CIENCIA POLÍTICA Y GOBIERNO": "11",
    "CONTABILIDAD": "12",
    "DERECHO": "13",
    "INGENIERÍA COMERCIAL": "14",
    "ARQUITECTURA": "15",
    "INGENIERÓA AGRONÓMICA Y AGRÍCOLA": "16",
    "INGENIERÍA AMBIENTAL": "17",
    "INGENIERÍA CIVIL": "18",
    "INGENIERÍA DE INDUSTRIA ALIMENTARIA": "19",
    "INGENIERÍA DE MINAS": "20",
    "INGENIERÍA DE SISTEMAS": "21",
    "INGENIERÍA ELECTRÓNICA": "22",
    "INGENIERÍA INDUSTRIAL": "23",
    "INGENIERÍA MECÁNICA": "24",
    "INGENIERÍA MECÁNICA ELÉCTRICA": "25",
    "INGENIERÍA MECATRÓNICA": "26",
    "MEDICINA VETERINARIA Y ZOOTECNIA": "27",
    "ENFERMERÍA": "28",
    "FARMACIA Y BIOQUÍMICA": "29",
    "INGENIERÍA BIOTECNOLÓGICA": "30",
    "MEDICINA HUMANA": "31",
    "OBSTRETICIA Y PUERICULTURA": "32",
    "ODONTOLOGÍA": "33",
    "TECNOLOGÍA MÉDICA": "34",
}


def conectar_db():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="123456",
        database="CertiGest"
    )

def exportar_asistentes(evento_CODIGO, evento_nombre):
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)

    # Obtener título, fecha e imagen del evento
    cursor.execute("SELECT TITULO, DATE_FORMAT(FECHA, '%d-%m-%Y') as FECHA, IMAGEN FROM EVENTO WHERE CODIGO = %s", (evento_CODIGO,))
    evento_info = cursor.fetchone()
    titulo = evento_info["TITULO"]
    fecha = evento_info["FECHA"]
    imagen_bytes = evento_info["IMAGEN"]

    # Obtener datos de asistentes
    cursor.execute("""
        SELECT a.DNI, a.APELLIDOS, a.NOMBRES, c.NOMBRE AS CARRERA, a.CORREO, a.SEMESTRE
        FROM ASISTENTE a
        JOIN CARRERA c ON a.ID_CARRERA = c.ID
        JOIN ASISTENCIA s ON s.DNI = a.DNI
        WHERE s.CODIGO_EVENTO = %s
    """, (evento_CODIGO,))
    rows = cursor.fetchall()
    df = pd.DataFrame(rows)

    if df.empty:
        messagebox.showinfo("Sin datos", "No hay asistentes para este evento.")
        return

    carpeta = f"Asistentes_{evento_nombre.replace(' ', '_')}"
    os.makedirs(carpeta, exist_ok=True)

    for carrera, grupo in df.groupby("CARRERA"):
        archivo = os.path.join(carpeta, f"{carrera}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Asistentes"

        # Título del evento en la parte superior
        ws.merge_cells('A1:F1')
        ws['A1'].value = f"{titulo} - {fecha}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Encabezados de columnas desde fila 3
        start_row = 3
        for col_index, col_name in enumerate(grupo.columns, start=1):
            cell = ws.cell(row=start_row, column=col_index, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Insertar datos desde la fila 4 en adelante
        for row_index, row in enumerate(grupo.values, start=start_row + 1):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

        # Ajustar ancho de columnas
        for col_index, column in enumerate(grupo.columns, start=1):
            max_length = max(len(str(cell)) for cell in grupo[column].values)
            adjusted_width = max(max_length + 2, 15)
            col_letter = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Insertar imagen después de la tabla
        if imagen_bytes:
            image_stream = io.BytesIO(imagen_bytes)
            img = ExcelImage(image_stream)
            img.width = 300
            img.height = 150
            final_row = start_row + len(grupo) + 3  # Espacio después de la tabla
            img_cell = f"A{final_row}"
            ws.add_image(img, img_cell)

        wb.save(archivo)

    messagebox.showinfo("Exportación Completa", f"Asistentes exportados en la carpeta: {carpeta}")
    conn.close()

def abrir_crear_evento():
    for widget in root.winfo_children():
        widget.destroy()

    frame = tk.Frame(root, bg="#f0f0f0")
    frame.pack(fill="both", expand=True)

    titulo_evento = tk.StringVar()
    imagen_path = tk.StringVar()
    excel_path = tk.StringVar()

    def seleccionar_imagen():
        filepath = filedialog.askopenfilename(title="Seleccionar imagen", filetypes=[("Archivos JPG", "*.jpg")])
        if filepath:
            imagen_path.set(filepath)
            lbl_imagen.config(text=f"Imagen seleccionada: {os.path.basename(filepath)}")

    def seleccionar_excel():
        filepath = filedialog.askopenfilename(title="Seleccionar archivo Excel", filetypes=[("Archivos Excel", "*.xlsx *.xls")])
        if filepath:
            excel_path.set(filepath)
            lbl_excel.config(text=f"Excel seleccionado: {os.path.basename(filepath)}")

    def guardar_evento():
        titulo = titulo_evento.get().strip()
        img_path = imagen_path.get()
        exc_path = excel_path.get()
        tipo = tipo_evento.get()

        if not titulo:
            messagebox.showwarning("Campo vacío", "Ingrese un título para el evento.")
            return
        if not img_path:
            messagebox.showwarning("Imagen faltante", "Seleccione una imagen para el evento.")
            return
        if not exc_path:
            messagebox.showwarning("Excel faltante", "Seleccione un archivo Excel con los asistentes.")
            return

        conn = conectar_db()
        cursor = conn.cursor()
        with open(img_path, "rb") as f:
            imagen_bytes = f.read()
        codigo_generado =cursor.callproc('insertar_evento', (titulo, imagen_bytes, tipo, ""))
        evento_codigo = codigo_generado[3] 

        try:
            df = pd.read_excel(exc_path)

            columnas_esperadas = {"DNI", "APELLIDOS", "NOMBRES", "ESCUELA PROFESIONAL", "E-MAIL", "SEMESTRE", "ASISTIO"}
            if not columnas_esperadas.issubset(df.columns):
                raise ValueError("El archivo Excel no contiene las columnas requeridas.")

            insertados = 0
            for _, row in df.iterrows():
                dni = str(row["DNI"]).strip()
                apellidos = str(row["APELLIDOS"]).strip().upper()
                nombres = str(row["NOMBRES"]).strip().upper()
                carrera_nombre = str(row["ESCUELA PROFESIONAL"]).strip().upper()
                correo = str(row["E-MAIL"]).strip()
                semestre = str(row["SEMESTRE"])
                asistencia = str(row["ASISTIO"]).strip().upper()
                if asistencia == "SI":
                    asistencia = 1
                else:
                    asistencia = 0

                id_carrera = carrera_map.get(carrera_nombre)
                if not id_carrera:
                    continue  

                cursor.execute("SELECT COUNT(*) FROM ASISTENTE WHERE DNI = %s", (dni,))
                existe = cursor.fetchone()[0]

                if not existe:
                    datos_asistente =(dni, apellidos, nombres, id_carrera, correo, semestre)
                    cursor.callproc('insertar_asistente', datos_asistente)
                
                datos_asistencia =(dni, evento_codigo, asistencia )
                cursor.callproc('insertar_asistencia', datos_asistencia)

                insertados += 1

            conn.commit()
            messagebox.showinfo("Éxito", f"Evento creado correctamente con ID: {evento_codigo}\n{insertados} registros procesados del Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar el archivo Excel: {str(e)}")

        
        conn.close()
        abrir_menu()
    estilo_boton1 = {"width": 20, "padx": 5, "pady": 10, "bg": "#4CAF50", "fg": "white", "font": ("Arial", 10, "bold")}

    tk.Label(frame, text="Título del Evento:", bg="#f0f0f0", font=("Arial", 12)).pack(pady=10)
    entry_titulo = tk.Entry(frame, textvariable=titulo_evento, font=("Arial", 12), width=40)
    entry_titulo.pack(pady=5)

    tipo_evento = tk.StringVar(value="EVENTO")
    tk.Label(frame, text="Tipo:", bg="#f0f0f0", font=("Arial", 12)).pack(pady=10)
    opciones = tk.OptionMenu(frame, tipo_evento, "EVENTO", "TALLER")
    opciones.config(
        font=("Arial", 14),
        bg="#4CAF50",  
        fg="white",
        width=20,
        highlightthickness=0,
        bd=0,
        activebackground="#4CAF50"
    )
    opciones["menu"].config(
        font=("Arial", 12),
        bg="#EDE7F6",
        fg="black"
    )
    opciones.pack(pady=5)
    

    tk.Label(frame, text="Imagen del Evento:", bg="#f0f0f0", font=("Arial", 12)).pack(pady=10)
    tk.Button(frame, text="Seleccionar Imagen", command=seleccionar_imagen, 
              **estilo_boton1).pack(pady=5)
    lbl_imagen = tk.Label(frame, text="Ninguna imagen seleccionada", bg="#f0f0f0")
    lbl_imagen.pack(pady=5)

    tk.Label(frame, text="Lista de Asistentes (Excel):", bg="#f0f0f0", font=("Arial", 12)).pack(pady=10)
    tk.Button(frame, text="Seleccionar Archivo Excel", command=seleccionar_excel, 
              **estilo_boton1).pack(pady=5)
    lbl_excel = tk.Label(frame, text="Ningún archivo seleccionado", bg="#f0f0f0")
    lbl_excel.pack(pady=5)

    frame_botones = tk.Frame(frame, bg="#f0f0f0")
    frame_botones.pack(pady=20)

    estilo_boton2 = {"width": 20, "padx": 5, "pady": 10, "bg": "#000000", "fg": "white", "font": ("Arial", 10, "bold")}

    tk.Button(frame_botones, text="Guardar Evento", command=guardar_evento, 
              **estilo_boton2).pack(side="left", padx=10)
    tk.Button(frame_botones, text="Volver al Menú", command=abrir_menu, 
              **estilo_boton2).pack(side="left", padx=10)

def abrir_consultas():
    for widget in root.winfo_children():
        widget.destroy()
    
    def cargar_eventos():
        conn = conectar_db()
        cursor = conn.cursor()
        cursor.callproc('cargar_eventos')
        for result in cursor.stored_results():
            eventos = result.fetchall()
        for ev in eventos:
            tree.insert("", "end", values=ev)
        conn.close()
    
    def buscar_eventos():
        query = busqueda_var.get().strip()
        criterio = criterio_var.get()

        for i in tree.get_children():
            tree.delete(i)

        if not query:
            cargar_eventos()
            return

        for i in tree.get_children():
            tree.delete(i)

        conn = conectar_db()
        cursor = conn.cursor()
            

        if criterio == "Código":
            cursor.execute("SELECT CODIGO, TITULO, DATE_FORMAT(FECHA, '%d-%m-%Y') FROM EVENTO WHERE CODIGO LIKE %s", (f"%{query}%",))
        elif criterio == "Título":
            cursor.execute("SELECT CODIGO, TITULO, DATE_FORMAT(FECHA, '%d-%m-%Y') FROM EVENTO WHERE TITULO LIKE %s", (f"%{query}%",))
        elif criterio == "Fecha":
            try:
                mes, anio = map(int, query.split("-"))  # ejemplo: "06-2025"    
                cursor.execute("""
                    SELECT CODIGO, TITULO, DATE_FORMAT(FECHA, '%d-%m-%Y') AS FECHA 
                    FROM EVENTO 
                    WHERE MONTH(FECHA) = %s AND YEAR(FECHA) = %s
                """, (mes, anio))
            except:
                messagebox.showerror("Error", "Ingrese mes y año en formato MM-AAAA. Ej: 06-2025")
                conn.close()
                return

        eventos = cursor.fetchall()
        for ev in eventos:
            tree.insert("", "end", values=ev)
        conn.close()

    def mostrar_inscritos():
        seleccionado = tree.selection()
        if not seleccionado:
            messagebox.showwarning("Seleccione un evento", "Debe seleccionar un evento.")
            return
        item = tree.item(seleccionado)
        evento_CODIGO, titulo, fecha = item['values']

        conn = conectar_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT a.DNI, a.APELLIDOS, a.NOMBRES, c.NOMBRE AS CARRERA, a.CORREO, a.SEMESTRE
            FROM ASISTENTE a
            JOIN CARRERA c ON a.ID_CARRERA = c.ID
            JOIN ASISTENCIA s ON s.DNI = a.DNI
            WHERE s.CODIGO_EVENTO = %s ORDER BY a.APELLIDOS
        """, (evento_CODIGO,))
        rows = cursor.fetchall()
        conn.close()

        if not rows:
            messagebox.showinfo("Sin datos", "No se encontraron registros.")
            return

        df = pd.DataFrame(rows)
        mostrar_datos(df)

    def mostrar_asistentes():
        seleccionado = tree.selection()
        if not seleccionado:
            messagebox.showwarning("Seleccione un evento", "Debe seleccionar un evento.")
            return
        item = tree.item(seleccionado)
        evento_CODIGO, titulo, fecha = item['values']

        conn = conectar_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT a.DNI, a.APELLIDOS, a.NOMBRES, c.NOMBRE AS CARRERA, a.CORREO, a.SEMESTRE
            FROM ASISTENTE a
            JOIN CARRERA c ON a.ID_CARRERA = c.ID
            JOIN ASISTENCIA s ON s.DNI = a.DNI
            WHERE s.CODIGO_EVENTO = %s AND ASISTIO = %s ORDER BY a.APELLIDOS    
        """, (evento_CODIGO, True,))
        rows = cursor.fetchall()
        conn.close()

        if not rows:
            messagebox.showinfo("Sin datos", "No se encontraron registros.")
            return

        df = pd.DataFrame(rows)
        mostrar_datos(df)

    def mostrar_datos(df):
        for widget in root.winfo_children():
            widget.destroy()

        frame_resultado = tk.Frame(root)
        frame_resultado.pack(fill="both", expand=True, padx=10, pady=10)

        # Crear Treeview con scroll
        tree_datos = ttk.Treeview(frame_resultado, show="headings")
        tree_datos.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frame_resultado, orient="vertical", command=tree_datos.yview)
        scrollbar.pack(side="right", fill="y")
        tree_datos.configure(yscroll=scrollbar.set)

        # Configurar columnas
        tree_datos["columns"] = list(df.columns)
        for col in df.columns:
            tree_datos.heading(col, text=col)
            tree_datos.column(col, width=100, anchor="center")

        # Insertar datos
        for _, row in df.iterrows():
            tree_datos.insert("", "end", values=list(row))

        # Botón volver
        tk.Button(root, text="Volver", command=abrir_consultas,
                bg="#FF5722", fg="white", font=("Arial", 10, "bold")).pack(pady=10)

    def exportar():
        seleccionado = tree.selection()
        if not seleccionado:
            messagebox.showwarning("Seleccione un evento", "Debe seleccionar un evento.")
            return
        item = tree.item(seleccionado)
        evento_CODIGO, titulo, fecha = item['values']
        exportar_asistentes(evento_CODIGO, titulo)

    def mostrar_imagen():
        seleccionado = tree.selection()
        if not seleccionado:
            messagebox.showwarning("Seleccione un evento", "Debe seleccionar un evento.")
            return

        item = tree.item(seleccionado)
        evento_CODIGO, titulo, fecha = item['values']

        conn = conectar_db()
        cursor = conn.cursor()
        cursor.execute("SELECT IMAGEN FROM EVENTO WHERE CODIGO = %s", (evento_CODIGO,))
        resultado = cursor.fetchone()
        conn.close()

        if not resultado or resultado[0] is None:
            messagebox.showinfo("Sin imagen", "No se encontró imagen para este evento en la base de datos.")
            return

        imagen_bytes = resultado[0]
        img = Image.open(io.BytesIO(imagen_bytes))
        img = img.resize((500, 400))
        img_tk = ImageTk.PhotoImage(img)

        for widget in root.winfo_children():
            widget.destroy()

        frame_img = tk.Frame(root)
        frame_img.pack(expand=True, fill="both")

        lbl = tk.Label(frame_img, image=img_tk)
        lbl.image = img_tk
        lbl.pack(pady=20)

        tk.Button(frame_img, text="Volver", command=abrir_consultas, bg="#4CAF50", fg="white", width=20, padx=5, pady=5, font=("Arial", 10, "bold")).pack(pady=10)

    estilo_boton = {"width": 20, "padx": 5, "pady": 5, "bg": "#4CAF50", "fg": "white", "font": ("Arial", 10, "bold")}
    frame_busqueda = tk.Frame(root)
    frame_busqueda.pack(pady=10)

    busqueda_var = tk.StringVar()
    criterio_var = tk.StringVar(value="Código")  # Por defecto
    criterios = ["Código", "Título", "Fecha"]


    tk.OptionMenu(frame_busqueda, criterio_var, *criterios).pack(side="left", padx=5)
    tk.Entry(frame_busqueda, textvariable=busqueda_var, font=("Arial", 10), width=40).pack(side="left", padx=5)
    tk.Button(frame_busqueda, text="Buscar", command=buscar_eventos, **estilo_boton).pack(side="left", padx=5)

    frame_superior = tk.Frame(root)
    frame_superior.pack(fill="both", expand=True)

    global tree
    tree = ttk.Treeview(frame_superior, columns=("Código", "Título", "Fecha"), show="headings")
    tree.heading("Código", text="Código")
    tree.heading("Título", text="Título")
    tree.heading("Fecha", text="Fecha de Creación")
    tree.column("Código", width=80, anchor="center")
    tree.column("Título", width=500)
    tree.column("Fecha", width=100, anchor="center")
    tree.pack(side="left", fill="both", expand=True, pady=10, padx=10)

    scrollbar = ttk.Scrollbar(frame_superior, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    

    frame_botones = tk.Frame(root)
    frame_botones.pack(pady=10)

    tk.Button(frame_botones, text="Ver Inscritos", command=mostrar_inscritos, **estilo_boton).pack(side="left")
    tk.Button(frame_botones, text="Ver Asistentes", command=mostrar_asistentes, **estilo_boton).pack(side="left")
    tk.Button(frame_botones, text="Exportar por Carrera", command=exportar, **estilo_boton).pack(side="left")
    tk.Button(frame_botones, text="Ver Imagen", command=mostrar_imagen, **estilo_boton).pack(side="left")
    tk.Button(frame_botones, text="Menú Principal", command=abrir_menu, **estilo_boton).pack(side="left")
    
    cargar_eventos()

def abrir_menu():
    for widget in root.winfo_children():
        widget.destroy()

    frame_inicio = tk.Frame(root, bg="#f0f0f0")
    frame_inicio.pack(expand=True)

    btn1 = tk.Button(
        frame_inicio, text="Crear Evento o Taller", width=30, height=3,
        bg="#2196F3", fg="white", font=("Arial", 16, "bold"),
        command=abrir_crear_evento
    )
    btn1.pack(pady=30)

    btn2 = tk.Button(
        frame_inicio, text="Consultas", width=30, height=3,
        bg="#4CAF50", fg="white", font=("Arial", 16, "bold"),
        command=abrir_consultas
    )
    btn2.pack(pady=10)

root = tk.Tk()
root.title("CertiGest")
root.geometry("1400x700")
root.configure(bg="#f0f0f0")
abrir_menu()
root.mainloop()